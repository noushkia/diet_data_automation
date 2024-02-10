"""
    Add new patient record to the database.
    The database is a directory of .docx files.
    Each file is a record of a patient.
    The file name format is current year and month + 3-digit id.
    For example, 9901001.docx
    The id is generated automatically.
    The file is generated from a template (template.docx).
    The template is a Word document with placeholders.
    The placeholders are replaced with the data from the form.
    The placeholders are the same as the keys in the context dictionary.
    The context dictionary is generated from the form.
    The form is a tkinter window.
    The form is generated from the char_data dictionary.
"""
import os
import pathlib
import re

from docxtpl import DocxTemplate
import datetime
from openpyxl import Workbook, load_workbook

from date.date import gregorian_to_jalali

RECORDS_PATH = "./db/patients/"
SUMMARIES_FILE = "patient_summaries.xlsx"
FORMAT = ".docx"
INITIAL_ID = "001"


def calculate_bmi(weight: float, height: float) -> float:
    """
    Calculate BMI given the weight and height.
    :param weight: weight in kg
    :param height: height in cm
    :return: BMI
    """
    return weight / ((height / 100) ** 2)


def add_patient_file(context):
    tpl = DocxTemplate("template.docx")
    bmi = calculate_bmi(float(context["bmi_weight"]), float(context["height"]))
    context["bmi"] = f"{bmi:.2f}"

    tpl.render(context)
    tpl.save(RECORDS_PATH + str(context["id"]) + FORMAT)


def add_patient_summary(context):
    # Store patient summaries in an Excel file
    summary_data = {
        "id": context["id"],
        "name": context["name"],
        "birthplace": context["birthplace"],
        "age": context["age"],
        "occupation": context["occupation"],
        "height": context["height"],
        "weight": context["weight"],
    }

    summary_file_path = RECORDS_PATH + SUMMARIES_FILE

    # Load the existing workbook or create a new one if it doesn't exist
    try:
        workbook = load_workbook(filename=summary_file_path)
    except FileNotFoundError:
        workbook = Workbook()

    # Get the worksheet or create a new one
    if "Sheet1" in workbook.sheetnames:
        worksheet = workbook["Sheet1"]
    else:
        worksheet = workbook.active
        worksheet.title = "Sheet1"

    # Create the column headers if the worksheet is empty
    if worksheet.max_row <= 1:
        headers = list(summary_data.keys())
        worksheet.append(headers)

    # Append the new data as a single row
    row_data = list(summary_data.values())
    worksheet.append(row_data)
    workbook.save(summary_file_path)


def generate_id():
    curr_date = gregorian_to_jalali(datetime.datetime.now())

    patients_files = [
        f for f in pathlib.Path(RECORDS_PATH).iterdir()
        if f.is_file() and re.match(fr"{str(curr_date.year)[2:4]}_{curr_date.month:02d}_\d{{3}}\.docx$", f.name)
    ]

    if not patients_files:
        return f"{str(curr_date.year)[2:4]}_{curr_date.month:02d}_{INITIAL_ID}"

    sorted_files = sorted(patients_files, key=lambda f: (
        int(f.stem.split("_")[2])
    ))

    latest_file = sorted_files[-1]

    try:
        last_year, last_month, last_id = latest_file.stem.split("_")
    except ValueError:  # .docx file name formats are invalid
        return f"{str(curr_date.year)[2:4]}_{curr_date.month:02d}_{INITIAL_ID}"

    # Check if the last record is from the current month
    if str(curr_date.year)[2:4] == last_year and f'{curr_date.month:02d}' == last_month:
        return f"{last_year}_{last_month}_{str(int(last_id) + 1).zfill(3)}"
    # If the last record is from another month
    else:
        return f"{str(curr_date.year)[2:4]}_{curr_date.month:02d}_{INITIAL_ID}"
